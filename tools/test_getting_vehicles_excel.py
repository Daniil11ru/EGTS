import argparse, re, sys, requests
from urllib.parse import urlencode, unquote
from openpyxl import load_workbook

EXPECTED_HEADERS = ["ID", "IMEI", "OID", "Название", "ID провайдера", "Статус модерации"]

def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--url", required=True)
    p.add_argument("--output", default=None)
    p.add_argument("--api-key", default=None)
    p.add_argument("--provider-id", type=int)
    p.add_argument("--moderation-status")
    p.add_argument("--timeout", type=int, default=30)
    p.add_argument("--save-only", action="store_true")
    p.add_argument("--allowed-status", default="pending,approved,rejected")
    return p.parse_args()

def pick_filename(res, fallback):
    cd = res.headers.get("content-disposition", "")
    m = re.search(r"filename\*=UTF-8''([^;]+)|filename=\"?([^\";]+)\"?", cd, re.I)
    if m:
        name = m.group(1) or m.group(2)
        try:
            return unquote(name)
        except Exception:
            return name
    return fallback or "vehicles.xlsx"

def is_int_like(v):
    if v is None: return False
    if isinstance(v, bool): return False
    if isinstance(v, int): return True
    if isinstance(v, float): return v.is_integer()
    if isinstance(v, str): return v.isdigit()
    return False

def validate_xlsx(path, allowed_statuses):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    header = [ws.cell(row=1, column=i+1).value for i in range(len(EXPECTED_HEADERS))]
    if header != EXPECTED_HEADERS:
        raise ValueError(f"неожиданный заголовок: {header} != {EXPECTED_HEADERS}")
    rows = ws.max_row - 1
    errors = []
    for r in range(2, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        d = ws.cell(row=r, column=4).value
        e = ws.cell(row=r, column=5).value
        f = ws.cell(row=r, column=6).value
        if not is_int_like(a): errors.append((r, "ID"))
        if b is None: errors.append((r, "IMEI"))
        if c is not None and not is_int_like(c): errors.append((r, "OID"))
        if e is None or not is_int_like(e): errors.append((r, "ID провайдера"))
        if f is None or str(f) not in allowed_statuses: errors.append((r, "Статус модерации"))
        if len(errors) >= 20: break
    if errors:
        preview = ", ".join([f"строка {r}: {col}" for r, col in errors[:10]])
        raise ValueError(f"ошибки в данных ({len(errors)}): {preview}")

def main():
    args = parse_args()
    params = {}
    if args.provider_id is not None: params["provider_id"] = str(args.provider_id)
    if args.moderation_status: params["moderation_status"] = args.moderation_status
    url = args.url if not params else (args.url + ("&" if "?" in args.url else "?") + urlencode(params))
    headers = {}
    if args.api_key: headers["X-API-Key"] = args.api_key
    with requests.get(url, headers=headers, timeout=args.timeout, stream=True) as res:
        if not res.ok:
            sys.stderr.write(f"HTTP {res.status_code}: {res.text[:500]}\n")
            sys.exit(1)
        filename = pick_filename(res, args.output)
        outpath = args.output or filename
        with open(outpath, "wb") as f:
            for chunk in res.iter_content(chunk_size=1024 * 128):
                if chunk: f.write(chunk)
    if args.save_only:
        print(f"Файл сохранён в {outpath}")
        sys.exit(0)
    allowed_statuses = {s.strip() for s in args.allowed_status.split(",") if s.strip()}
    validate_xlsx(outpath, allowed_statuses)
    print(f"Файл сохранён в {outpath}")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        sys.stderr.write(f"Ошибка: {e}\n")
        sys.exit(2)
